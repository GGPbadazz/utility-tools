"""
P&ID Device & Pipe Extractor  (DXF mode via ezdxf)
===================================================
Reads a clean DXF file and extracts:
  - Devices : INSERT (block references) with block name, layer, insert point,
              and all ATTRIB values (tag numbers, descriptions, etc.)
  - Pipes   : LINE, LWPOLYLINE, POLYLINE, ARC, SPLINE on pipe-related layers

Output: pid_extract.xlsx + pid_pipes.csv + pid_devices.csv
"""

import argparse
import sys, csv, math, re
from pathlib import Path
from collections import Counter

import ezdxf
from ezdxf import recover

WORK_DIR  = Path(__file__).parent
DXF_FILE  = WORK_DIR / "sample_pid.dxf"
OUT_XLSX  = WORK_DIR / "pid_extract.xlsx"
OUT_PIPES = WORK_DIR / "pid_pipes.csv"
OUT_DEVS  = WORK_DIR / "pid_devices.csv"
OUT_TAGS  = WORK_DIR / "pid_pipe_tags.csv"

PIPE_LAYER_KW = ["pipe","管","管道","line","process","工艺","pid","实线","阀","介质"]
PIPE_TYPES    = {"LINE","LWPOLYLINE","POLYLINE","ARC","SPLINE"}

# Pipe tag pattern, e.g. CW-P01001-50-CS, HW-P02001-100-CS-H50, CA-P03001-15-SS
PIPE_TAG_RE = re.compile(r"\b[A-Z]{2,5}-P\d{3,6}[A-Z]?\d?-\d{1,4}-[A-Z0-9]{2,6}(?:-[A-Z0-9]{2,6})?\b", re.I)

def parse_args():
    parser = argparse.ArgumentParser(description="Extract devices, pipe entities, and pipe tags from a P&ID DXF.")
    parser.add_argument("dxf", nargs="?", default=str(DXF_FILE), help="Path to the DXF file to read.")
    parser.add_argument("--out-dir", default=str(WORK_DIR), help="Directory for CSV/XLSX outputs.")
    return parser.parse_args()

def clean_mtext(s):
    s = re.sub(r"\\[A-Za-z][^;]*;", "", s or "")
    return s.replace("\\P", " ").replace("{", "").replace("}", "").strip()

def is_pipe_layer(name):
    n = (name or "").lower()
    return any(kw.lower() in n for kw in PIPE_LAYER_KW)

def entity_length(e):
    t = e.dxftype()
    try:
        if t == "LINE":
            s, end = e.dxf.start, e.dxf.end
            return math.hypot(end.x - s.x, end.y - s.y)
        if t == "LWPOLYLINE":
            pts = [(p[0], p[1]) for p in e.get_points("xy")]
            return sum(math.hypot(pts[i+1][0]-pts[i][0], pts[i+1][1]-pts[i][1])
                       for i in range(len(pts)-1))
        if t == "POLYLINE":
            verts = [(v.dxf.location.x, v.dxf.location.y) for v in e.vertices]
            return sum(math.hypot(verts[i+1][0]-verts[i][0], verts[i+1][1]-verts[i][1])
                       for i in range(len(verts)-1))
        if t == "ARC":
            r = e.dxf.radius
            sa = math.radians(e.dxf.start_angle)
            ea = math.radians(e.dxf.end_angle)
            span = (ea - sa) % (2*math.pi)
            return r * span
        if t == "SPLINE" and hasattr(e, "approx_length"):
            return e.approx_length(num=100)
    except Exception:
        pass
    return 0.0

def entity_endpoints(e):
    t = e.dxftype()
    try:
        if t == "LINE":
            return (e.dxf.start.x, e.dxf.start.y, e.dxf.end.x, e.dxf.end.y)
        if t == "LWPOLYLINE":
            pts = list(e.get_points("xy"))
            if pts:
                return (pts[0][0], pts[0][1], pts[-1][0], pts[-1][1])
        if t == "POLYLINE":
            verts = [(v.dxf.location.x, v.dxf.location.y) for v in e.vertices]
            if verts:
                return (verts[0][0], verts[0][1], verts[-1][0], verts[-1][1])
        if t == "ARC":
            return (e.dxf.center.x, e.dxf.center.y, "", "")
    except Exception:
        pass
    return ("", "", "", "")

def extract_attribs(insert):
    attrs = {}
    try:
        for a in insert.attribs:
            tag = (a.dxf.tag or "").strip().upper()
            val = (a.dxf.text or "").strip()
            attrs[tag] = val
    except Exception:
        pass
    return attrs

def load_doc(path):
    if not path.exists():
        print(f"ERROR: {path} not found.")
        sys.exit(1)
    print(f"Loading {path.name}  ({path.stat().st_size//1024} KB) ...")
    doc, auditor = recover.readfile(str(path))
    if auditor.has_errors:
        print(f"  {len(auditor.errors)} audit errors (recovered)")
    return doc

def extract(doc):
    pipes, devices, tags = [], [], []
    layers = set()
    msp = doc.modelspace()
    for e in msp:
        t = e.dxftype()
        layer = e.dxf.layer if e.dxf.hasattr("layer") else "0"
        layers.add(layer)

        if t in ("TEXT", "MTEXT"):
            try:
                raw = e.dxf.text if t == "TEXT" else e.text
            except Exception:
                raw = ""
            txt = clean_mtext(raw)
            for m in PIPE_TAG_RE.finditer(txt):
                tag = m.group(0).upper()
                parts = tag.split("-")
                try:
                    ip = e.dxf.insert
                    x, y = round(ip.x, 3), round(ip.y, 3)
                except Exception:
                    x = y = ""
                tags.append({
                    "pipe_tag":   tag,
                    "service":    parts[0] if len(parts) > 0 else "",
                    "line_no":    parts[1] if len(parts) > 1 else "",
                    "size":       parts[2] if len(parts) > 2 else "",
                    "spec":       parts[3] if len(parts) > 3 else "",
                    "insulation": parts[4] if len(parts) > 4 else "",
                    "layer":      layer,
                    "x":          x,
                    "y":          y,
                    "full_text":  txt,
                })

        if t == "INSERT":
            attrs = extract_attribs(e)
            devices.append({
                "block_name":  e.dxf.name,
                "layer":       layer,
                "x":           round(e.dxf.insert.x, 3),
                "y":           round(e.dxf.insert.y, 3),
                "rotation":    round(getattr(e.dxf, "rotation", 0), 2),
                "tag":         attrs.get("TAG", attrs.get("TAGNO", attrs.get("ITEM",""))),
                "description": attrs.get("DESC", attrs.get("DESCRIPTION", attrs.get("SERVICE",""))),
                "size":        attrs.get("SIZE", attrs.get("LINESIZE","")),
                "all_attrs":   "; ".join(f"{k}={v}" for k,v in attrs.items()),
            })
        elif t in PIPE_TYPES:
            length = entity_length(e)
            xs, ys, xe, ye = entity_endpoints(e)
            pipes.append({
                "entity":      t,
                "layer":       layer,
                "likely_pipe": is_pipe_layer(layer),
                "length":      round(length, 3),
                "x_start":     round(xs, 3) if xs != "" else "",
                "y_start":     round(ys, 3) if ys != "" else "",
                "x_end":       round(xe, 3) if xe != "" else "",
                "y_end":       round(ye, 3) if ye != "" else "",
            })
    return pipes, devices, tags, sorted(layers)

def print_summary(pipes, devices, tags, layers):
    likely = [p for p in pipes if p["likely_pipe"]]
    uniq_tags = sorted({t["pipe_tag"] for t in tags})
    print(f"\n{'='*60}\nEXTRACTION SUMMARY\n{'='*60}")
    print(f"Total layers       : {len(layers)}")
    print(f"Total INSERT blocks: {len(devices)}")
    print(f"Total line entities: {len(pipes)}")
    print(f"Likely pipe lines  : {len(likely)}")
    print(f"Pipe tag instances : {len(tags)}")
    print(f"Unique pipe tags   : {len(uniq_tags)}")
    svc = Counter(t["service"] for t in tags)
    print("\nPipe-tags by service prefix:")
    for k, v in svc.most_common():
        print(f"  {v:5d}x  {k}")
    print("\nFirst 40 unique pipe tags:")
    for tg in uniq_tags[:40]:
        print(f"  {tg}")
    bc = Counter(d["block_name"] for d in devices)
    print("\nTop 25 blocks:")
    for name, cnt in bc.most_common(25):
        if not name.startswith("*"):
            print(f"  {cnt:5d}x  {name}")
    pl = sorted({p["layer"] for p in likely})
    print(f"\nPipe layers ({len(pl)}):")
    for l in pl: print(f"  {l}")
    print(f"\nAll layers ({len(layers)}):")
    for n in layers: print(f"  {n}")

def save_csv(rows, path):
    if not rows: return
    with open(path,"w",newline="",encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)
    print(f"CSV: {path}  ({len(rows)} rows)")

def save_excel(pipes, devices, tags, layers):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed - Excel skipped"); return
    wb = openpyxl.Workbook()
    def sheet(title, rows, color, first=False):
        ws = wb.active if first else wb.create_sheet(title)
        ws.title = title
        if not rows: ws.append(["No data"]); return
        hdrs = list(rows[0].keys())
        ws.append(hdrs)
        for ci,_ in enumerate(hdrs,1):
            c = ws.cell(1,ci); c.font=Font(bold=True,color="FFFFFF")
            c.fill=PatternFill("solid",fgColor=color); c.alignment=Alignment(horizontal="center")
        for r in rows: ws.append(list(r.values()))
        for ci in range(1,len(hdrs)+1):
            mw = max(len(str(ws.cell(r,ci).value or "")) for r in range(1,ws.max_row+1))
            ws.column_dimensions[get_column_letter(ci)].width = min(mw+4,60)
    sheet("Devices",      devices,                              "4472C4", first=True)
    sheet("Pipe Tags",    tags,                                 "C00000")
    sheet("All Lines",    pipes,                                "70AD47")
    sheet("Likely Pipes", [p for p in pipes if p["likely_pipe"]],"ED7D31")
    ws2 = wb.create_sheet("Layers"); ws2.title="Layers"; ws2.append(["Layer Name"])
    for n in layers: ws2.append([n])
    wb.save(str(OUT_XLSX)); print(f"Excel: {OUT_XLSX}")

if __name__ == "__main__":
    args = parse_args()
    out_dir = Path(args.out_dir).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)
    DXF_FILE = Path(args.dxf).expanduser()
    OUT_XLSX  = out_dir / "pid_extract.xlsx"
    OUT_PIPES = out_dir / "pid_pipes.csv"
    OUT_DEVS  = out_dir / "pid_devices.csv"
    OUT_TAGS  = out_dir / "pid_pipe_tags.csv"

    doc = load_doc(DXF_FILE)
    pipes, devices, tags, layers = extract(doc)
    print_summary(pipes, devices, tags, layers)
    save_csv(pipes, OUT_PIPES)
    save_csv(devices, OUT_DEVS)
    save_csv(tags, OUT_TAGS)
    save_excel(pipes, devices, tags, layers)
    print("\nDone.")
